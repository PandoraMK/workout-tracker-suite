import datetime
import json
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
import streamlit as st

# --- Load User Profile First (Required for st.set_page_config) ---
user_profile_file = "user_profile.json"
default_profile = {
    "name": "Modiri",
    "body_weight": 88.0,
    "gender": "Male",
    "age": 25,
    "height": 178.0,
}

if os.path.exists(user_profile_file):
  try:
    with open(user_profile_file, "r") as f:
      loaded_profile = json.load(f)
      default_profile.update(loaded_profile)
  except:
    pass

current_user = default_profile.get("name", "Modiri")

st.set_page_config(
    page_title=f"{current_user}'s Workout Master Suite",
    page_icon="💪",
    layout="centered",
)

file_path = "Workout_Master_Suite.xlsx"
reviews_file_path = "Workout_Reviews.xlsx"


def create_default_workbook(path):
  wb = Workbook()
  ws = wb.active
  ws.title = "Workout Log"
  ws.append([])
  ws.append([])
  ws.append([])
  headers = [
      "Date",
      "Routine / Focus",
      "Exercise",
      "Sets",
      "Reps",
      "Weight (kg)",
      "Total Volume (kg)",
      "RPE (1-10)",
  ]
  ws.append(headers)
  wb.save(path)


def create_default_reviews_workbook(path):
  wb = Workbook()
  ws = wb.active
  ws.title = "Reviews"
  headers = [
      "Date",
      "Time",
      "Tester Name",
      "Rating (1-5)",
      "Category",
      "Feedback Message",
  ]
  ws.append(headers)
  wb.save(path)


if not os.path.exists(file_path):
  create_default_workbook(file_path)

if not os.path.exists(reviews_file_path):
  create_default_reviews_workbook(reviews_file_path)

if "user_profile" not in st.session_state:
  st.session_state.user_profile = default_profile

# Sidebar for User Profile & Settings (Persistent)
with st.sidebar:
  st.markdown("### ⚙️ Athlete Profile")
  p = st.session_state.user_profile

  entered_name = st.text_input("Name", value=p.get("name", "Modiri"))
  entered_bw = st.number_input(
      "Body Weight (kg)",
      min_value=30.0,
      max_value=250.0,
      value=float(p.get("body_weight", 88.0)),
      step=0.5,
  )
  entered_gender = st.selectbox(
      "Gender",
      ["Male", "Female", "Other"],
      index=(
          0
          if p.get("gender", "Male") == "Male"
          else (1 if p.get("gender") == "Female" else 2)
      ),
  )
  entered_age = st.number_input(
      "Age", min_value=10, max_value=100, value=int(p.get("age", 25))
  )
  entered_height = st.number_input(
      "Height (cm)",
      min_value=100.0,
      max_value=250.0,
      value=float(p.get("height", 178.0)),
      step=1.0,
  )

  if st.button("Save Profile"):
    st.session_state.user_profile = {
        "name": entered_name,
        "body_weight": entered_bw,
        "gender": entered_gender,
        "age": entered_age,
        "height": entered_height,
    }
    try:
      with open(user_profile_file, "w") as f:
        json.dump(st.session_state.user_profile, f)
    except:
      pass
    st.success("Profile saved! Refresh page to update title.")

current_user = st.session_state.user_profile.get("name", "Modiri")

st.title(f"💪 {current_user}'s Workout Master Suite")
st.write(
    f"Elite training tracker for **{current_user}** (BW:"
    f" {st.session_state.user_profile.get('body_weight')}kg) built with home"
    " workouts, cardio tracking, & separate reviews logging."
)

# Navigation Tabs for enhanced structure
tab1, tab2, tab3, tab4 = st.tabs(
    [
        "📝 Logger Form",
        "📈 Progress & Analytics",
        "📖 Glossary & Definitions",
        "💬 Feedback & Reviews",
    ]
)

with tab1:
  st.subheader("Add Exercise / Activity Details")

  routine_exercises_map = {
      "Upper Body A": [
          "Barbell Bench Press",
          "Incline DB Press",
          "Overhead Shoulder Press",
          "Cable Lateral Raises",
          "DB Lateral Raises",
          "Tricep Pushdowns",
      ],
      "Upper Body B": [
          "Lat Pulldown",
          "Seated Cable Row",
          "Neutral Grip Pull-Ups",
          "Barbell Bent-Over Row",
          "Face Pulls",
          "Rear Delt Fly",
      ],
      "Lower Body A": [
          "Barbell Back Squat",
          "Leg Press",
          "Romanian Deadlift",
          "Standing Calf Raises",
      ],
      "Lower Body B": [
          "Bulgarian Split Squat",
          "Goblet Squat",
          "Romanian Deadlift",
          "Leg Press",
      ],
      "Home Workouts": [
          "Standard Push-Ups",
          "Pike Push-Ups",
          "Bodyweight Squats",
          "Chair Bulgarian Split Squats",
          "Walking Lunges",
          "Glute Bridges",
          "Chair Dips",
          "Superman Back Extensions",
          "Plank Hold",
      ],
      "Cardio": [
          "Outside Running",
          "Indoor Treadmill Run",
          "Assault Bike",
          "Rowing Machine",
      ],
      "Full Body": [
          "Barbell Back Squat",
          "Barbell Bench Press",
          "Lat Pulldown",
          "Standing Calf Raises",
      ],
  }

  col1, col2 = st.columns(2)
  with col1:
    routine_options = list(routine_exercises_map.keys()) + ["Custom"]
    routine = st.selectbox("Routine / Focus", routine_options)

    if routine == "Custom":
      routine_name = st.text_input("Enter custom routine name", "Custom Focus")
      available_exercises = [
          "Barbell Back Squat",
          "Barbell Bench Press",
          "Lat Pulldown",
      ]
    else:
      routine_name = routine
      available_exercises = routine_exercises_map.get(routine, [])

  with col2:
    exercise_choice = st.selectbox(
        "Select Exercise / Activity", available_exercises + ["Other (Type Below)"]
    )
    if exercise_choice == "Other (Type Below)":
      exercise_name = st.text_input("Type Exercise Name", "New Exercise")
    else:
      exercise_name = exercise_choice

  # --- TARGETED STRETCHING GUIDE BASED ON ROUTINE ---
  st.markdown("---")
  with st.expander(
      f"🧘 View Recommended Stretches for: **{routine_name}**", expanded=False
  ):
    if "Upper" in routine_name:
      st.markdown("""
            * **Pre-Workout Dynamic Stretches (Before you lift):**
              * *Arm Circles:* 15 circles forward and backward to lubricate shoulder joints.
              * *Band Pull-Aparts or Doorway Chest Stretch:* 2 sets of 15 reps to activate upper back and open chest.
              * *Torso Twists:* 10 per side to mobilize thoracic spine.
            * **Post-Workout Static Stretches (After your workout):**
              * *Cross-Body Shoulder Stretch:* Hold 30 seconds per arm.
              * *Overhead Tricep Stretch:* Hold 30 seconds per arm.
              * *Doorway Chest Stretch:* Hold 30 seconds to relieve front delts and pectorals.
            """)
    elif "Lower" in routine_name:
      st.markdown("""
            * **Pre-Workout Dynamic Stretches (Before you lift):**
              * *Leg Swings:* 15 front-to-back and 15 side-to-side per leg.
              * *World's Greatest Stretch:* 5 reps per side (deep lunges with thoracic rotation).
              * *Bodyweight Squat Pries:* Hold bottom of a squat for 5 seconds x 5 reps.
            * **Post-Workout Static Stretches (After your workout):**
              * *Standing Quad Stretch:* Hold 30 seconds per leg.
              * *Seated Hamstring Stretch:* Hold 45 seconds focusing on breathing.
              * *Figure-4 Glute Stretch:* Hold 30 seconds per leg to release tight glutes/hips.
            """)
    elif "Cardio" in routine_name:
      st.markdown("""
            * **Pre-Workout Dynamic Stretches (Before you run/cardio):**
              * *High Knees & Butt Kicks:* 30 seconds each to elevate heart rate.
              * *Walking Lunges with Twist:* 10 steps per leg to open hip flexors.
              * *Ankle Bounces:* 20 reps to prep calves and Achilles tendons.
            * **Post-Workout Static Stretches (After your session):**
              * *Wall Calf Stretch:* Hold 45 seconds per leg.
              * *Hip Flexor Kneeling Stretch:* Hold 45 seconds per side.
              * *Hamstring Floor Stretch:* Hold 45 seconds.
            """)
    elif "Home Workouts" in routine_name:
      st.markdown("""
            * **Pre-Workout Dynamic Stretches (Before bodyweight exercises):**
              * *Arm Swings & Jumping Jacks:* 1 minute to warm up the whole body.
              * *Inchworms:* 5-8 reps to stretch hamstrings and activate shoulders/core.
              * *Dynamic Lunges:* 10 reps per leg.
            * **Post-Workout Static Stretches (After your home session):**
              * *Child’s Pose:* Hold 60 seconds to decompress lower back and lats.
              * *Downward Dog:* Hold 45 seconds for calves, hamstrings, and shoulders.
              * *Chest Opener Stretch:* Hold 30 seconds.
            """)
    else:
      st.markdown("""
            * **Pre-Workout Dynamic Stretches:** Arm circles, leg swings, light bodyweight movements.
            * **Post-Workout Static Stretches:** Full body static stretches focusing on worked muscle groups.
            """)

  st.markdown("---")

  # --- DYNAMIC FORM: CARDIO VS WEIGHTS ---
  if routine == "Cardio":
    st.write("🏃 **Cardio Metrics**")
    c1, c2, c3 = st.columns(3)
    with c1:
      cardio_dist = st.number_input(
          "Distance (km)", min_value=0.1, max_value=100.0, value=5.0, step=0.1
      )
    with c2:
      cardio_time = st.number_input(
          "Duration (mins)", min_value=1, max_value=600, value=30, step=1
      )
    with c3:
      cardio_hr = st.number_input(
          "Avg Heart Rate (bpm)", min_value=0, max_value=220, value=145, step=1
      )

    if cardio_dist > 0:
      pace_mins = int(cardio_time // cardio_dist)
      pace_secs = int(((cardio_time / cardio_dist) - pace_mins) * 60)
      pace_str = f"{pace_mins}m {pace_secs}s / km"
    else:
      pace_str = "N/A"

    st.info(
        f"📊 **Cardio Summary:** Distance: **{cardio_dist} km** | Duration:"
        f" **{cardio_time} mins** | Pace: **{pace_str}** | HR:"
        f" **{cardio_hr if cardio_hr > 0 else 'N/A'} bpm**"
    )

    total_sets = 1
    representative_reps = cardio_time
    weight_str = f"{cardio_dist} km"
    total_volume = 0
    total_volume_str = f"HR: {cardio_hr}bpm" if cardio_hr > 0 else "N/A"

  else:
    st.write("🏋️ **Set & Weight Progression (Pyramids / Weight Changes)**")

    num_blocks = st.selectbox(
        "How many different weight blocks?",
        [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
        index=0,
    )

    total_sets = 0
    total_volume = 0
    weight_parts = []
    representative_reps = 8

    for i in range(num_blocks):
      if num_blocks > 1:
        st.caption(f"Block {i+1}")

      c1, c2, c3 = st.columns(3)
      with c1:
        s = st.number_input(
            f"Sets ({i+1})",
            min_value=1,
            max_value=20,
            value=2 if i > 0 else 4,
            key=f"sets_{i}",
        )
      with c2:
        r = st.number_input(
            f"Reps ({i+1})",
            min_value=1,
            max_value=100,
            value=12 if routine == "Home Workouts" else 8,
            key=f"reps_{i}",
        )
      with c3:
        w = st.number_input(
            f"Weight kg ({i+1}) (0 for Bodyweight)",
            min_value=0.0,
            max_value=500.0,
            value=0.0 if routine == "Home Workouts" else 40.0 + (i * 5.0),
            step=2.5,
            key=f"weight_{i}",
        )

      total_sets += s
      if i == 0:
        representative_reps = r

      total_volume += s * r * w
      weight_parts.append(f"{w}kg")

    if len(weight_parts) == 1:
      weight_str = weight_parts[0]
    elif len(weight_parts) == 2:
      weight_str = f"{weight_parts[0]} & {weight_parts[1]}"
    else:
      weight_str = ", ".join(weight_parts[:-1]) + f" & {weight_parts[-1]}"

    total_volume_str = f"{total_volume}kg"

    st.info(
        f"📊 **Calculated Summary:** Total Sets: **{total_sets}** | Combined"
        f" Weight: **{weight_str}** | Total Volume: **{total_volume} kg**"
    )

  st.markdown("---")
  col_rpe, col_date = st.columns(2)
  with col_rpe:
    rpe = st.slider(
        "RPE (1-10)", min_value=1.0, max_value=10.0, value=8.0, step=0.5
    )
  with col_date:
    log_date = st.date_input("Workout Date", datetime.date.today())

  if st.button("Save Structured Entry to Excel"):
    try:
      wb = load_workbook(file_path)
      if "Workout Log" in wb.sheetnames:
        ws = wb["Workout Log"]
        next_row = ws.max_row + 1

        date_str = log_date.strftime("%Y/%m/%d")

        ws.cell(row=next_row, column=1, value=date_str)
        ws.cell(row=next_row, column=2, value=routine_name)
        ws.cell(row=next_row, column=3, value=exercise_name)
        ws.cell(row=next_row, column=4, value=total_sets)
        ws.cell(row=next_row, column=5, value=representative_reps)
        ws.cell(row=next_row, column=6, value=weight_str)
        ws.cell(row=next_row, column=7, value=total_volume_str)
        ws.cell(row=next_row, column=8, value=rpe)

        wb.save(file_path)
        st.success(
            f"Successfully logged {exercise_name} ({weight_str}) under"
            f" {routine_name}!"
        )
      else:
        st.error("'Workout Log' sheet not found.")
    except Exception as e:
      st.error(f"Error saving to Excel: {e}")

  st.markdown("---")
  st.subheader("📊 Live Workout Log Preview")
  try:
    if os.path.exists(file_path):
      df_log = pd.read_excel(file_path, sheet_name="Workout Log", skiprows=3)
      if not df_log.empty:
        st.dataframe(df_log.tail(6), use_container_width=True)
      else:
        st.info("No workout entries logged yet.")
  except Exception as e:
    st.info(f"Could not load log preview: {e}")

with tab2:
  st.subheader("📈 Training Analytics & Progress Charts")
  try:
    if os.path.exists(file_path):
      df_analytics = pd.read_excel(
          file_path, sheet_name="Workout Log", skiprows=3
      )
      if not df_analytics.empty and "Total Volume (kg)" in df_analytics.columns:
        df_analytics["Clean_Volume"] = (
            df_analytics["Total Volume (kg)"]
            .astype(str)
            .str.replace("kg", "", regex=False)
            .str.replace("HR: ", "", regex=False)
            .str.replace("bpm", "", regex=False)
        )
        df_analytics["Clean_Volume"] = pd.to_numeric(
            df_analytics["Clean_Volume"], errors="coerce"
        ).fillna(0)

        st.markdown("### Total Lift Volume / Activity Output Over Time")
        chart_data = df_analytics[["Date", "Clean_Volume"]].dropna()
        if not chart_data.empty:
          st.line_chart(
              chart_data.set_index("Date"), use_container_width=True
          )
        else:
          st.info("Log a few workouts to see your progression chart!")

        st.markdown("---")
        st.markdown("### Filter History by Routine / Focus")
        df_analytics["Routine / Focus"] = df_analytics[
            "Routine / Focus"
        ].fillna("Unassigned")
        unique_routines = sorted(
            df_analytics["Routine / Focus"].unique().tolist()
        )
        selected_routine = st.selectbox(
            "Select Routine to Inspect", unique_routines, key="analytics_routine"
        )
        filtered_df = df_analytics[
            df_analytics["Routine / Focus"] == selected_routine
        ]
        st.dataframe(filtered_df, use_container_width=True)
      else:
        st.info("Add entries to generate performance charts.")
    else:
      st.info("Workout file not found.")
  except Exception as e:
    st.info(f"Error loading charts: {e}")

with tab3:
  st.subheader("📖 Fitness Glossary & Terminology Guide")
  st.markdown("""
    * **RPE (Rate of Perceived Exertion):** A scale from **1 to 10** used to measure how intense a set felt relative to failure.
      * *10:* Absolute maximum effort (0 reps left in the tank).
      * *8 - 9:* Hard set; you could have realistically done 1 or 2 more reps.
      * *6 - 7:* Moderate effort; comfortable with several reps left.
    * **Total Volume:** The overall workload calculated as $\text{Sets} \times \text{Reps} \times \text{Weight}$. Tracking volume over time is key for progressive overload and muscle hypertrophy.
    * **Cardio Pace:** Calculated as time divided by distance ($\text{Minutes} / \text{km}$) to track running/cardio efficiency over time.
    * **Home Workouts:** Bodyweight training focused on time-under-tension, high rep ranges, and calisthenics progression.
    * **Routine / Focus:** The structural split of your training day (e.g., Upper/Lower Body, Home Workouts, Cardio) ensuring balanced recovery and growth.
    """)

with tab4:
  st.subheader("💬 Tester Suggestion & Review Box")
  st.write(
      "Have a suggestion, found a bug, or want to leave feedback on the app?"
      " Drop it below!"
  )

  with st.form("feedback_form"):
    fb_name = st.text_input("Your Name / Handle", value=current_user)
    fb_rating = st.slider(
        "Rating (1 = Needs Work, 5 = Awesome!)",
        min_value=1,
        max_value=5,
        value=5,
    )
    fb_category = st.selectbox(
        "Feedback Category",
        [
            "General Review",
            "Bug Report",
            "Feature Request",
            "UI / Design Suggestion",
        ],
    )
    fb_message = st.text_area(
        "Your Feedback or Suggestion",
        placeholder="Type your feedback here...",
    )
    submit_fb = st.form_submit_button("Submit Feedback")

    if submit_fb:
      if fb_message.strip():
        try:
          if not os.path.exists(reviews_file_path):
            create_default_reviews_workbook(reviews_file_path)

          wb_rev = load_workbook(reviews_file_path)
          ws_rev = wb_rev.active

          now_dt = datetime.datetime.now()
          fb_date_str = now_dt.strftime("%Y/%m/%d")
          fb_time_str = now_dt.strftime("%H:%M:%S")

          ws_rev.append(
              [
                  fb_date_str,
                  fb_time_str,
                  fb_name,
                  fb_rating,
                  fb_category,
                  fb_message,
              ]
          )
          wb_rev.save(reviews_file_path)
          st.success(
              "Thank you! Your feedback has been successfully saved to"
              " Workout_Reviews.xlsx."
          )
        except Exception as e:
          st.error(f"Error saving feedback: {e}")
      else:
        st.warning("Please type a message before submitting.")

  st.markdown("---")
  st.subheader("📥 Received Feedback & Reviews")
  try:
    if os.path.exists(reviews_file_path):
      df_fb = pd.read_excel(reviews_file_path)
      if not df_fb.empty:
        st.dataframe(df_fb, use_container_width=True)
      else:
        st.info("No reviews submitted yet.")
    else:
      st.info("No reviews file found.")
  except Exception as e:
    st.info(f"Could not load reviews: {e}")
